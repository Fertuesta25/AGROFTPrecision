# -*- coding: utf-8 -*-
"""
Núcleo de cálculo de uniformidad de riego por aspersión (modo matriz regular).
Solo depende de numpy (incluido en QGIS). Sin scipy.

Reproduce los coeficientes de Catch3D:
  - CU (Christiansen)
  - DU (uniformidad del cuarto inferior, criterio floor n/4)
  - Traslape por superposición periódica de la matriz
  - Conversión a pluviometría (mm/h) y balance de masa
"""
import numpy as np


def cu_christiansen(values):
    v = np.asarray(values, dtype=float).ravel()
    s = v.sum()
    if s <= 0:
        return 0.0
    m = v.mean()
    return float(100.0 * (1.0 - np.sum(np.abs(v - m)) / s))


def du_low_quarter(values):
    v = np.sort(np.asarray(values, dtype=float).ravel())
    n = len(v)
    if n == 0 or v.mean() == 0:
        return 0.0
    k = max(1, n // 4)                       # criterio Catch3D (floor)
    return float(100.0 * v[:k].mean() / v.mean())


def overlap_matrix(G, a, b):
    """Superposición periódica: marco de a filas x b columnas (en celdas)."""
    G = np.asarray(G, dtype=float)
    out = np.zeros((a, b), dtype=float)
    nr, nc = G.shape
    for r in range(nr):
        for c in range(nc):
            out[r % a, c % b] += G[r, c]
    return out


def unoverlapped_stats(G):
    """CU/DU del patrón individual, sobre los vasos con agua (>0), como Catch3D."""
    G = np.asarray(G, dtype=float)
    wet = G[G > 0]
    if wet.size == 0:
        return 0.0, 0.0
    return cu_christiansen(wet), du_low_quarter(wet)


def overlapped_stats(G, a, b):
    o = overlap_matrix(G, a, b)
    return cu_christiansen(o), du_low_quarter(o), o


def pluviometry_mm_h(mean_ml, diam_cm, duration_h):
    """Lámina (mm/h) a partir del volumen medio (ml), diámetro de boca y duración."""
    if diam_cm <= 0 or duration_h <= 0:
        return 0.0
    area_cm2 = np.pi * (diam_cm ** 2) / 4.0
    depth_mm = mean_ml * 10.0 / area_cm2
    return float(depth_mm / duration_h)


def mass_balance(pluv_measured_mm_h, Q_lph, Sl_m, Sm_m):
    """Pluviometría teórica = Q / area_marco y desviación relativa."""
    area = Sl_m * Sm_m
    if area <= 0 or Q_lph <= 0:
        return None, None
    teorica = Q_lph / area          # 1 L/m2 = 1 mm
    dev = pluv_measured_mm_h / teorica - 1.0 if teorica > 0 else None
    return float(teorica), float(dev)


def edge_warning(G):
    """Avisa si el patrón no se cierra a cero en los bordes (captación parcial)."""
    G = np.asarray(G, dtype=float)
    tot = G.sum()
    if tot <= 0:
        return False, {}
    border = {
        'fila_superior': G[0, :].sum(),
        'fila_inferior': G[-1, :].sum(),
        'col_izquierda': G[:, 0].sum(),
        'col_derecha': G[:, -1].sum(),
    }
    # umbral: 5 % del total en cualquier borde sugiere truncamiento
    thr = 0.05 * tot
    flag = any(val > thr for val in border.values())
    pct = {k: (100.0 * val / tot) for k, val in border.items()}
    return flag, pct


# ============================================================
#  Tabla comparativa de varios marcos
# ============================================================
def compare_marcos(G, marcos):
    """marcos: lista de (a, b) en celdas. Devuelve lista de dicts ordenada por CU desc."""
    rows = []
    for (a, b) in marcos:
        if a > G.shape[0] or b > G.shape[1] or a < 1 or b < 1:
            continue
        cu, du, o = overlapped_stats(G, a, b)
        rows.append({'a': a, 'b': b, 'cu': cu, 'du': du, 'mean': float(o.mean())})
    rows.sort(key=lambda r: r['cu'], reverse=True)
    return rows


def valoracion(cu):
    if cu >= 88: return "Excelente"
    if cu >= 84: return "Muy bueno"
    if cu >= 80: return "Bueno"
    if cu >= 75: return "Aceptable"
    return "Deficiente"


# ============================================================
#  Modo puntos GPS: construir matriz regular desde puntos
# ============================================================
def matrix_from_points(coords, values, spacing=None, snap_tol=0.5):
    """
    Convierte puntos georreferenciados a una matriz regular.
      coords: lista [(x, y), ...]   (coordenadas proyectadas, en metros)
      values: lista de volúmenes captados (ml), mismo orden
      spacing: tamaño de celda (m). Si None, se estima del vecino más cercano.
      snap_tol: tolerancia (fracción de celda) para ajustar puntos a la grilla.
    Devuelve (G, info) donde info trae spacing, filas, columnas, origen y nº de puntos sin ubicar.
    """
    xs = np.asarray([c[0] for c in coords], dtype=float)
    ys = np.asarray([c[1] for c in coords], dtype=float)
    vals = np.asarray(values, dtype=float)
    if len(xs) < 4:
        raise ValueError("Se requieren al menos 4 puntos.")

    if spacing is None or spacing <= 0:
        spacing = _estimate_spacing(xs, ys)
    if spacing <= 0:
        raise ValueError("No se pudo estimar el espaciamiento; indíquelo manualmente.")

    x0, y0 = xs.min(), ys.min()
    # índice de columna (x) y fila (y); fila 0 = y máxima (arriba)
    cols = np.rint((xs - x0) / spacing).astype(int)
    rows_from_bottom = np.rint((ys - y0) / spacing).astype(int)
    ncols = cols.max() + 1
    nrows = rows_from_bottom.max() + 1
    rows = (nrows - 1) - rows_from_bottom         # fila 0 arriba

    G = np.zeros((nrows, ncols), dtype=float)
    counts = np.zeros((nrows, ncols), dtype=int)
    unplaced = 0
    # control de ajuste a grilla
    dx = np.abs((xs - x0) / spacing - cols)
    dy = np.abs((ys - y0) / spacing - rows_from_bottom)
    for k in range(len(vals)):
        if dx[k] > snap_tol or dy[k] > snap_tol:
            unplaced += 1
        G[rows[k], cols[k]] += vals[k]
        counts[rows[k], cols[k]] += 1
    # promediar celdas con más de un punto
    mask = counts > 1
    G[mask] = G[mask] / counts[mask]
    info = {'spacing': float(spacing), 'nrows': int(nrows), 'ncols': int(ncols),
            'x0': float(x0), 'y0': float(y0), 'off_grid': int(unplaced)}
    return G, info


def _estimate_spacing(xs, ys):
    """Estima el espaciamiento como la mediana de la distancia al vecino más cercano."""
    n = len(xs)
    pts = np.column_stack([xs, ys])
    nn = []
    for i in range(n):
        d = np.sqrt((pts[:, 0] - pts[i, 0]) ** 2 + (pts[:, 1] - pts[i, 1]) ** 2)
        d[i] = np.inf
        nn.append(d.min())
    return float(np.median(nn))


# ============================================================
#  Rampas de color (clásicas de QGIS) -> (r,g,b) en [0,255]
# ============================================================
# nodos de control por rampa; interpolación lineal entre ellos
_RAMPS = {
    "RdYlGn":   [(165,0,38),(215,48,39),(244,109,67),(253,174,97),(254,224,139),
                 (217,239,139),(166,217,106),(102,189,99),(26,152,80),(0,104,55)],
    "Spectral": [(158,1,66),(213,62,79),(244,109,67),(253,174,97),(254,224,139),
                 (230,245,152),(171,221,164),(102,194,165),(50,136,189),(94,79,162)],
    "Viridis":  [(68,1,84),(72,40,120),(62,74,137),(49,104,142),(38,130,142),
                 (31,158,137),(53,183,121),(110,206,88),(181,222,43),(253,231,37)],
    "Blues":    [(247,251,255),(222,235,247),(198,219,239),(158,202,225),(107,174,214),
                 (66,146,198),(33,113,181),(8,81,156),(8,48,107)],
}

def ramp_names():
    return list(_RAMPS.keys())

def ramp_color(name, t):
    """t en [0,1] -> (r,g,b). Para RdYlGn/Spectral el bajo=rojo, alto=verde/azul."""
    nodes = _RAMPS.get(name, _RAMPS["RdYlGn"])
    t = max(0.0, min(1.0, float(t)))
    if t <= 0: return nodes[0]
    if t >= 1: return nodes[-1]
    pos = t * (len(nodes) - 1)
    i = int(pos); f = pos - i
    a = nodes[i]; b = nodes[i + 1]
    return (int(a[0]+(b[0]-a[0])*f), int(a[1]+(b[1]-a[1])*f), int(a[2]+(b[2]-a[2])*f))


# ============================================================
#  Lectura de matriz desde Excel (.xlsx) — matriz tal cual
# ============================================================
def _isnum(x):
    if x is None or x == "":
        return False
    try:
        float(str(x).replace(",", ".")); return True
    except (TypeError, ValueError):
        return False

def _mostly_text(seq):
    seq = [v for v in seq if v not in (None, "")]
    if not seq:
        return False
    return sum(0 if _isnum(v) else 1 for v in seq) > len(seq) / 2

def _looks_like_index(seq):
    nums = []
    for v in seq:
        if v in (None, ""):
            continue
        try:
            nums.append(float(str(v).replace(",", ".")))
        except (TypeError, ValueError):
            return False
    if len(nums) < 3:
        return False
    diffs = [nums[i + 1] - nums[i] for i in range(len(nums) - 1)]
    return all(abs(d - 1) < 1e-9 for d in diffs) or all(abs(d + 1) < 1e-9 for d in diffs)

def matrix_from_xlsx(path, sheet=None, has_header=None, has_labels=None):
    """
    Lee una matriz de volúmenes (ml) de Excel, formato filas x columnas (tal cual).
    has_header / has_labels: True/False para forzar; None = autodetectar
    (celda-esquina no numérica o fila/columna de índices consecutivos).
    Celdas vacías o de texto = 0. Devuelve (G, info).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("Falta la librería 'openpyxl'.")
    import numpy as _np
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    raw = [list(row) for row in ws.iter_rows(values_only=True)]
    raw = [r for r in raw if any(v not in (None, "") for v in r)]
    if not raw:
        raise ValueError("La hoja está vacía.")
    ncol = max(len(r) for r in raw)
    raw = [list(r) + [None] * (ncol - len(r)) for r in raw]

    corner_text = (raw[0][0] is None) or (not _isnum(raw[0][0]))
    header_body = [raw[0][k] for k in range(1, ncol)]
    auto_header = corner_text or _mostly_text(raw[0]) or _looks_like_index(header_body)
    drop_header = auto_header if has_header is None else has_header
    if drop_header:
        raw = raw[1:]
    col0 = [r[0] for r in raw]
    auto_label = corner_text or _mostly_text(col0) or _looks_like_index(col0)
    drop_label = auto_label if has_labels is None else has_labels
    if drop_label:
        raw = [r[1:] for r in raw]

    G = [[float(str(v).replace(",", ".")) if _isnum(v) else 0.0 for v in r] for r in raw]
    G = _np.array(G, dtype=float)
    if G.size == 0 or G.shape[0] < 2 or G.shape[1] < 2:
        raise ValueError("La matriz resultante es demasiado pequeña; revise encabezados.")
    info = {'nrows': int(G.shape[0]), 'ncols': int(G.shape[1]),
            'header_removed': bool(drop_header), 'label_removed': bool(drop_label),
            'sheet': ws.title}
    return G, info


def _build_matrix(raw, has_header=None, has_labels=None):
    import numpy as _np
    raw = [list(r) for r in raw if any(v not in (None, "") for v in r)]
    if not raw:
        raise ValueError("Sin datos.")
    ncol = max(len(r) for r in raw)
    raw = [list(r) + [None] * (ncol - len(r)) for r in raw]
    corner_text = (raw[0][0] is None) or (not _isnum(raw[0][0]))
    header_body = [raw[0][k] for k in range(1, ncol)]
    auto_h = corner_text or _mostly_text(raw[0]) or _looks_like_index(header_body)
    drop_h = auto_h if has_header is None else has_header
    if drop_h:
        raw = raw[1:]
    col0 = [r[0] for r in raw]
    auto_l = corner_text or _mostly_text(col0) or _looks_like_index(col0)
    drop_l = auto_l if has_labels is None else has_labels
    if drop_l:
        raw = [r[1:] for r in raw]
    G = [[float(str(v).replace(",", ".")) if _isnum(v) else 0.0 for v in r] for r in raw]
    G = _np.array(G, dtype=float)
    if G.size == 0 or G.shape[0] < 2 or G.shape[1] < 2:
        raise ValueError("La matriz resultante es demasiado pequeña; revise encabezados.")
    return G, bool(drop_h), bool(drop_l)


def xlsx_sheets(path):
    from openpyxl import load_workbook
    return load_workbook(path, read_only=True).sheetnames


def matrix_from_csv(path, has_header=None, has_labels=None):
    import csv
    with open(path, newline='', encoding='utf-8-sig') as f:
        sample = f.read(2048); f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        raw = [row for row in csv.reader(f, dialect)]
    G, dh, dl = _build_matrix(raw, has_header, has_labels)
    return G, {'nrows': int(G.shape[0]), 'ncols': int(G.shape[1]),
               'header_removed': dh, 'label_removed': dl, 'sheet': 'CSV'}


def build_html_report(ctx):
    """Genera el reporte HTML (auto-contenido) con mapa de color. Devuelve str."""
    import numpy as _np
    ramp = ctx.get('ramp', 'RdYlGn')

    def color_table(M, dec=0):
        M = _np.asarray(M, float)
        mn, mx = M.min(), M.max()
        rng = (mx - mn) if mx > mn else 1.0
        out = ['<table class="grid"><tr><th></th>']
        out.append(''.join('<th>%d</th>' % (j + 1) for j in range(M.shape[1])) + '</tr>')
        for i in range(M.shape[0]):
            out.append('<tr><th>%d</th>' % (i + 1))
            for j in range(M.shape[1]):
                r, g, b = ramp_color(ramp, (M[i, j] - mn) / rng)
                lum = 0.299 * r + 0.587 * g + 0.114 * b
                fg = '#000' if lum > 140 else '#fff'
                out.append('<td style="background:rgb(%d,%d,%d);color:%s">%.*f</td>'
                           % (r, g, b, fg, dec, M[i, j]))
            out.append('</tr>')
        out.append('</table>')
        return ''.join(out)

    p = ctx['params']; sp = ctx['spacing']
    cmp_html = ''
    if ctx.get('comparison'):
        cmp_html = ('<h2>Comparacion de marcos</h2><table class="tbl">'
                    '<tr><th>Marco (m)</th><th>CU (%)</th><th>DU (%)</th><th>Valoracion</th></tr>')
        for r in ctx['comparison']:
            cmp_html += ('<tr><td>%.0f &times; %.0f</td><td>%.1f</td><td>%.1f</td><td>%s</td></tr>'
                         % (r['a'] * sp, r['b'] * sp, r['cu'], r['du'], valoracion(r['cu'])))
        cmp_html += '</table>'

    dev = ctx.get('dev')
    dev_html = ('%+.1f %%' % (dev * 100)) if dev is not None else '&mdash;'
    edge_html = ''
    if ctx.get('edge_flag'):
        pk = ctx.get('edge_peak', ('', 0))
        edge_html = ('<p class="warn">&#9888; Aviso: el patron no se cierra a cero en el borde (%s = %.1f %% '
                     'del total). Posible captacion parcial.</p>' % (pk[0].replace('_', ' '), pk[1]))

    css = ('body{font-family:Calibri,Arial,sans-serif;color:#222;margin:24px;}'
           'h1{color:#1F4E5F;border-bottom:2px solid #1F4E5F;padding-bottom:4px;}'
           'h2{color:#1F4E5F;margin-top:22px;}table{border-collapse:collapse;margin:8px 0;}'
           '.tbl td,.tbl th{border:1px solid #bbb;padding:4px 10px;font-size:13px;}'
           '.tbl th{background:#1F4E5F;color:#fff;}'
           '.grid td,.grid th{border:1px solid #ccc;padding:3px 6px;text-align:center;font-size:11px;min-width:22px;}'
           '.grid th{background:#1F4E5F;color:#fff;}'
           '.kpi{display:inline-block;background:#E8F0E3;border:1px solid #cdddc8;border-radius:6px;'
           'padding:8px 16px;margin:4px;font-size:15px;}.kpi b{color:#1F6B2F;font-size:20px;}'
           '.warn{color:#c00000;font-weight:bold;}.meta{color:#555;font-size:13px;}')

    h = []
    h.append('<!doctype html><html><head><meta charset="utf-8"><style>%s</style></head><body>' % css)
    h.append('<h1>Evaluacion de uniformidad de riego por aspersion</h1>')
    h.append('<p class="meta">%s</p>' % ctx.get('subtitle', ''))
    h.append('<h2>Resultados principales &mdash; marco %s</h2>' % ctx['marco_str'])
    h.append('<div class="kpi">CU (traslapado)<br><b>%.1f %%</b></div>' % ctx['cu'])
    h.append('<div class="kpi">DU (traslapado)<br><b>%.1f %%</b></div>' % ctx['du'])
    h.append('<div class="kpi">Pluviometria media<br><b>%.1f mm/h</b></div>' % ctx['pluv'])
    h.append('<p>Patron individual (sin traslape): CU = %.1f %% &middot; DU = %.1f %%<br>'
             'Pluviometria: min %.1f &middot; max %.1f mm/h &middot; Balance de masa: teorica %.1f mm/h '
             '&middot; desviacion %s</p>' % (ctx['cu0'], ctx['du0'], ctx['pluv_min'],
                                             ctx['pluv_max'], ctx.get('teorica', 0), dev_html))
    h.append(edge_html)
    h.append('<h2>Parametros de campo</h2><table class="tbl">'
             '<tr><th>Parametro</th><th>Valor</th><th>Parametro</th><th>Valor</th></tr>'
             '<tr><td>Aspersor (fila, columna)</td><td>%s, %s</td><td>Altura del aspersor</td><td>%.2f m</td></tr>'
             '<tr><td>Presion de trabajo</td><td>%.0f kPa</td><td>Caudal</td><td>%.0f L/h</td></tr>'
             '<tr><td>Velocidad del viento</td><td>%.1f m/s</td><td>Direccion del viento</td><td>%.0f&deg;</td></tr>'
             '<tr><td>Duracion del ensayo</td><td>%.0f min</td><td>Diametro del pluviometro</td><td>%.1f cm</td></tr>'
             '<tr><td>Espaciamiento de la malla</td><td>%.2f m</td><td>Marco de riego</td><td>%s</td></tr></table>'
             % (p['asp_row'], p['asp_col'], p['altura'], p['presion'], p['caudal'], p['viento'],
                p['vientodir'], p['tiempo'], p['diam'], sp, ctx['marco_str']))
    h.append('<h2>Matriz de volumenes captados (ml)</h2>')
    h.append(color_table(ctx['input_matrix'], 0))
    h.append('<h2>Matriz traslapada &mdash; lamina relativa (rampa: %s)</h2>' % ramp)
    h.append(color_table(ctx['overlap_matrix'], 0))
    h.append(cmp_html)
    h.append('<p class="meta" style="margin-top:24px">Generado por el plugin &laquo;Uniformidad de Aspersion&raquo; '
             '&middot; superposicion periodica (Catch3D).</p></body></html>')
    return ''.join(h)


def point_to_cell(x, y, info):
    """Mapea una coordenada (x,y) a la celda (fila, columna) 1-indexada de la matriz
    reconstruida, usando el mismo origen/espaciamiento que matrix_from_points."""
    sp = info['spacing']; x0 = info['x0']; y0 = info['y0']
    nrows = info['nrows']; ncols = info['ncols']
    col = int(round((x - x0) / sp))
    rfb = int(round((y - y0) / sp))           # fila desde abajo
    row = (nrows - 1) - rfb                    # fila 0 = arriba
    col = max(0, min(ncols - 1, col))
    row = max(0, min(nrows - 1, row))
    return row + 1, col + 1
